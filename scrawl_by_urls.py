# encoding:utf-8
# Author Yundong xiao

# ############################Import Libs
# standard lib
import time

# third party lib
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import ErrorInResponseException
from selenium.common.exceptions import NoSuchElementException
from openpyxl import Workbook
from openpyxl import load_workbook

# self written lib
#
# ###########################Import Libs End

# ###########################Global Data
# output file name
Fetch_Result_File = "./file/fetch_result.xlsx"
DEBUG = True
RECODING_PER_TIME = 100
ERROR_BLUE_V = "BLUE V"
ERROR_UNREGISTER = "UNREGISTER"
ERROR_LOAD_FANS_PAGE = "UNABLE TO LOAD FANS PAGE"
MAIN_PAGE = "https://weibo.com/"
USERNAME = "18771038375"
PASSWORD = "xyd123456"
driver = None
# ###########################Global Data End
# ###########################Functions
# login


def login(page_url=MAIN_PAGE, username=USERNAME, password=PASSWORD, try_times=5):
    # login in page setting
    # 18771038375 18602710227
    global driver
    while True:
        driver = webdriver.Chrome()
        driver.implicitly_wait(5)
        driver.maximize_window()
        driver.set_page_load_timeout(120)
        driver.delete_all_cookies()
        try:
            driver.get(page_url)
            try_times -= 1
            WebDriverWait(driver, 30).until(lambda x: x.find_element_by_xpath('//*[@id="pl_login_form"]/div/div[1]/\
            div/a[1]'))
            driver.find_element_by_xpath('//*[@id="loginname"]').send_keys(username)
            driver.find_element_by_xpath('//*[@id="pl_login_form"]/div/div[3]/div[2]/div/input').send_keys(password)
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="pl_login_form"]/div/div[3]/div[6]/a').send_keys(Keys.ENTER)
            WebDriverWait(driver, 30).until(lambda x: x.find_element_by_xpath('//*[@id="v6_pl_rightmod_myinfo"]/div/div\
            /div[2]/div/a[1]'))
            break
        except TimeoutException:
            if try_times == 0:
                driver.quit()
                print "Unable to Login in!", try_times, "times"
                return None
            print "Unable to Login in!"
            driver.quit()
            time.sleep(1)
    print "Successful Login in!"
    return driver

# try_to_load_page and handler exception carefully


def load_page(page_url, try_times=5):
    global driver
    while True:
        try:
            if try_times == 0:
                break
            else:
                driver.get(page_url)
            return True
        except TimeoutException:
            print "TimeoutException"
            try_times -= 1
            driver.quit()
            time.sleep(2)
            driver = login()
            time.sleep(2)
        except ErrorInResponseException:
            print "ErrorInResponseException"
            try_times -= 1
            time.sleep(180)
    if try_times == 0:
        if DEBUG is True:
            while True:
                time.sleep(60)
        else:
            print "Network Disable!!!"
            return False

# main scrawl function recursion itself if there is a deep url needed to be scrawled


def scrawl(url, finished_list, error_list):
    if load_page(url) is False:
        error_list.append([url, ERROR_LOAD_FANS_PAGE])
        return
    else:
        try:
            driver.find_elements_by_xpath('//*[starts-with(@id,"Pl_Official")]/div/\
    div/div/div[2]/div[1]/ul/li[1]/dl/dd[1]/div[2]/span')
        except NoSuchElementException:
            try:
                driver.find_element_by_xpath('//*[starts-with(@id,"Pl_Official_")]/div[1]/div/div[2]/div[2]/h1')
                print "No fans data :", url
                return
            except NoSuchElementException:
                print "Impossible Case in scrawl:", url
                if DEBUG is True:
                    while True:
                        time.sleep(60)
    # get all data
    try:
        xpath = '//*[@id="Pl_Official_Headerv6__1"]/div[1]/div/div[2]/div[2]/h1'
        nickname = driver.find_element_by_xpath(xpath)
        xpath = '//*[starts-with(@id,"Pl_Official_")]/div/div/div/div[2]/div[1]/ul/*/dl/dd[1]/div[1]/a[1]\
        [starts-with(@usercard,"id=")]'
        list_fans_names_and_urls = driver.find_elements_by_xpath(xpath)
        xpath = '//*[starts-with(@id,"Pl_Official_")]/div/div/div/div[2]/div[1]/ul/*/dl/dd[1]/div[2]/span[1]/em/a'
        list_fans_concerns = driver.find_elements_by_xpath(xpath)
        xpath = '//*[starts-with(@id,"Pl_Official_")]/div/div/div/div[2]/div[1]/ul/*/dl/dd[1]/div[2]/span[2]/em/a'
        list_fans_fans = driver.find_elements_by_xpath(xpath)
        xpath = '//*[starts-with(@id,"Pl_Official_")]/div/div/div/div[2]/div[1]/ul/*/dl/dd[1]/div[2]/span[3]/em/a'
        list_fans_weibo = driver.find_elements_by_xpath(xpath)
    except NoSuchElementException:
        print "Impossible Case NoSuchElementException:", url
        if DEBUG is True:
            while True:
                time.sleep(60)
        return
    try:
        # extract
        for i in range(len(list_fans_names_and_urls)):
            finished_list.append([nickname.text, list_fans_names_and_urls[i].text, list_fans_names_and_urls[i].
                                 get_attribute("href"), list_fans_concerns[i].text, list_fans_fans[i].text,
                                  list_fans_weibo[i].text])
    except IndexError:
        print "IndexError"
        print len(list_fans_names_and_urls), len(list_fans_concerns)
        for i in range(len(list_fans_names_and_urls)):
            print list_fans_names_and_urls[i].text
        for i in range(len(list_fans_concerns)):
            print list_fans_concerns[i].text

        if DEBUG is True:
            while True:
                time.sleep(60)

    if url.find("page=5") >= 0:
        return
    # go to next page to recurse
    try:
        next_page = driver.find_element_by_link_text('下一页').get_attribute("href")
        if next_page is None:
            print "No next page: ", url
            return
        time.sleep(1)
        scrawl(next_page, finished_list, error_list)
        return
    except NoSuchElementException:
        print "No next page: ", url
        return
# prepare_scrawl get fan's link then enter it


def prepare_scrawl(url, return_list, error_list):
    if load_page(url) is False:
        error_list.append([url, ERROR_LOAD_FANS_PAGE])
        return
    else:
        try:
            WebDriverWait(driver, 5).until(lambda x: x.find_element_by_xpath('//*[@id="v6_pl_rightmod_myinfo"]/\
            div/div/div[2]/div/a[1]'))
            print "unregister url: ", url
            error_list.append([url, ERROR_UNREGISTER])
            return
        except TimeoutException:
            pass
    # go to fans page
    try:
        scrawl_url_fans_page = (WebDriverWait(driver, 5).until(lambda x: x.find_elements_by_xpath('//*[@class\
         ="t_link S_txt1"]')))[1].get_attribute("href")
    except TimeoutException:
            print "blue v no fans link", url
            error_list.append([url, ERROR_BLUE_V])
            # if DEBUG is True:
            #    while True:
            #        time.sleep(60)
            return
    # exist_element = (driver, '//*[@id="Pl_Core_T8CustomTriColumn__3"]/div/div/div/table/tbody/tr/td[2]/a')
    # scrawl start from this page
    scrawl(scrawl_url_fans_page, return_list, error_list)
    print "Result Numbers :", len(result_list)

# ###########################Functions End


if __name__ == '__main__':

    driver = login()
    # Recording time start
    print time.strftime("%Y-%m-%d %H:%M %p", time.localtime())

    # collecting pages start
    workbook_read = load_workbook('./file/urls.xlsx')
    sheets_read = workbook_read.sheetnames
    sheet_read = workbook_read[sheets_read[0]]

    urls = []
    rows = sheet_read.rows
    for row in sheet_read.rows:
        for cell in row:
            urls.append(cell.value)

    # open write unfinished file
    workbook_write = Workbook()
    sheet_write_result = workbook_write.create_sheet(title="results")
    sheet_write_failed = workbook_write.create_sheet(title="failed urls")

    # For URL in URL_List
    idx_file = 0
    for scrawl_url in urls:
        idx_file += 1
        if idx_file % RECODING_PER_TIME == 0:
            workbook_write.save(Fetch_Result_File)  # append result and failed case into execl RECODING_PER_TIME
            print time.strftime("%Y-%m-%d %H:%M %p", time.localtime()), "idx: ", idx_file

        result_list = []
        exception_list = []
        scrawl_url = scrawl_url[0:scrawl_url.find("refer_flag")]
        print "URL Count: ", idx_file, scrawl_url
        prepare_scrawl(scrawl_url, result_list, exception_list)
        # write to sheet_write
        for element in result_list:
            sheet_write_result.append(element)
        for element in exception_list:
            sheet_write_failed.append(element)
    workbook_write.save(Fetch_Result_File)

    # Recording end time
    print time.strftime("%Y-%m-%d %H:%M %p", time.localtime())
    print "Scrawl End"
