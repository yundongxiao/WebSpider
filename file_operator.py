# encoding:utf-8
from openpyxl import Workbook
from openpyxl import load_workbook

# open workbook , then get write sheet , append data , save in the end
workbook_write = Workbook()
sheet_write = workbook_write.active
sheet_write.append(["中文", "うくぐ"])
sheet_write.append(["네이버", "ภาษาไทย"])
sheet_write.append(["English", "123"])
workbook_write.save("./file/test_write_execl.xlsx")
sheet_write.append(["中文", "うくぐ"])
sheet_write.append(["네이버", "ภาษาไทย"])
sheet_write.append(["English", "123"])
workbook_write.save("./file/test_write_execl.xlsx")

# open workbook , then get first sheet to print
workbook_read = load_workbook('./file/test_write_execl.xlsx')
sheets = workbook_read.sheetnames
# get first sheet to print
sheet_print = workbook_read[sheets[0]]
rows = sheet_print.rows
for row in sheet_print.rows:
    for cell in row:
        print (cell.value)
